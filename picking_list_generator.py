"""
Générateur automatique de Picking Lists pour la fabrication de blades d'éoliennes
Gère le stock en FIFO et génère les picking lists avec emplacements
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.page import PageMargins
from datetime import datetime
from pathlib import Path
import copy
import re
import shutil
import json

_ARCHIVE_DIR        = Path('picking_lists_archive')
_ARCHIVE_INDEX_FILE = _ARCHIVE_DIR / 'archive_index.json'
_PL_COUNTER_FILE    = Path('pl_counter.json')


def _get_next_pl_number(shared_dir=None):
    """Retourne le prochain numéro de picking list (partagé si shared_dir fourni)."""
    counter_file = (Path(shared_dir) / 'pl_counter.json') if shared_dir else _PL_COUNTER_FILE
    counter = {'next_number': 1}
    if counter_file.exists():
        try:
            counter = json.loads(counter_file.read_text(encoding='utf-8'))
        except Exception:
            pass
    num = counter['next_number']
    counter['next_number'] = num + 1
    counter_file.parent.mkdir(parents=True, exist_ok=True)
    counter_file.write_text(json.dumps(counter, indent=2), encoding='utf-8')
    return num

def sort_location_key(location):
    """
    Crée une clé de tri pour un emplacement
    
    Format numérique: "17-0605/3" → (0, 17, 605, 3)
    Format texte: "QUAI51-PCW" → (1, "QUAI51-PCW")
    
    Args:
        location: Emplacement sous forme de chaîne
    
    Returns:
        tuple: Clé de tri
    """
    if not location or location in ['MANQUANT', 'NON DISPONIBLE']:
        return (2, location)  # Mettre les manquants à la fin
    
    # Pattern pour format numérique: XX-XXXX/X
    pattern = r'^(\d+)-(\d+)/(\d+)$'
    match = re.match(pattern, str(location))
    
    if match:
        allee = int(match.group(1))
        travee = int(match.group(2))
        niveau = int(match.group(3))
        return (0, allee, travee, niveau)  # 0 = numérique (en premier)
    else:
        return (1, str(location))  # 1 = texte (après les numériques)


class PickingListGenerator:
    def __init__(self, excel_file_path):
        """
        Initialise le générateur avec le fichier Excel
        
        Args:
            excel_file_path: Chemin vers le fichier Excel principal
        """
        self.excel_file = excel_file_path
        self.stock_df = None
        self.components = {}
        self.picking_lists = {}

    def _generate_stock_cache(self, sap_df):
        """Génère stock_cache.xlsx : version allégée du stock pour consultation rapide.
        Appelé automatiquement après chaque import SAP."""
        cache_cols = [
            'Product',
            'Storage Bin',
            'Quantity',
            'Handling Unit',
            'Shelf Life Expiration Date',
        ]
        # Inclure la colonne description SAP si elle existe (nom quelconque contenant 'description')
        desc_col = next(
            (c for c in sap_df.columns if 'description' in str(c).lower()),
            None
        )
        if desc_col:
            cache_cols.append(desc_col)

        # Inclure la colonne unité SAP si elle existe (contient 'unit' mais pas 'handling')
        unit_col = next(
            (c for c in sap_df.columns
             if 'unit' in str(c).lower() and 'handling' not in str(c).lower()),
            None
        )
        if unit_col:
            cache_cols.append(unit_col)

        available = [c for c in cache_cols if c in sap_df.columns]
        cache_df = sap_df[available].copy()
        if 'Quantity' in cache_df.columns:
            cache_df['Quantity'] = pd.to_numeric(cache_df['Quantity'], errors='coerce')
        dst = Path(self.excel_file).parent / 'stock_cache.xlsx'
        cache_df.to_excel(dst, index=False, engine='openpyxl')
        desc_info = f" + colonne '{desc_col}'" if desc_col else ''
        print(f"   📦 Cache stock : {dst.name} ({len(cache_df)} lignes{desc_info})")

    def _backup_excel(self, file_path=None, max_backups=5):
        """Crée une copie horodatée de main.xlsx dans backups/ avant toute écriture."""
        src = Path(file_path or self.excel_file)
        if not src.exists():
            return
        backup_dir = src.parent / 'backups'
        backup_dir.mkdir(exist_ok=True)
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        dst = backup_dir / f'{src.stem}_{ts}{src.suffix}'
        shutil.copy2(src, dst)
        # Garder uniquement les N backups les plus récents
        backups = sorted(backup_dir.glob(f'{src.stem}_*{src.suffix}'))
        for old in backups[:-max_backups]:
            old.unlink()
        print(f"   💾 Backup : {dst.name}")


    def import_stock_from_sap(self, file_path):
        """
        Remplace l'onglet Stock de main.xlsx avec les données de l'export SAP.

        Args:
            file_path: Chemin vers le fichier SAP export (.xlsx / .xls)
        Returns:
            bool: True si l'import a réussi
        """
        print("\n📥 Import du stock SAP...")

        sap_path = Path(file_path)
        print(f"   → Fichier sélectionné : {sap_path.name}")

        # Lire entièrement en str pour préserver la précision des grands entiers (HU…)
        # pandas/float64 ne peut représenter que ~15 chiffres significatifs,
        # ce qui tronque silencieusement les identifiants à 18+ chiffres.
        sap_df = pd.read_excel(sap_path, dtype=str)
        print(f"   ✓ {len(sap_df)} lignes lues")

        # Vérifier la présence des colonnes obligatoires
        required_cols = [
            'Product',
            'Quantity',
            'Storage Bin',
            'Handling Unit',
            'Goods Receipt Date',
            'Shelf Life Expiration Date',
        ]
        missing = [c for c in required_cols if c not in sap_df.columns]
        if missing:
            raise ValueError(
                f"Colonnes manquantes dans le fichier SAP :\n"
                + "\n".join(f"   • {c}" for c in missing)
                + f"\n\nColonnes présentes : {list(sap_df.columns)}"
            )

        def _to_cell(v):
            """Convertit une valeur str en type Excel adapté.
            Grands entiers (≥ 1e14) → texte pour éviter la troncature des HU."""
            if not isinstance(v, str) or v.lower() in ('nan', 'nat', 'none', ''):
                return None
            try:
                f = float(v)
                if abs(f) >= 1e14 and f == int(f):
                    return v          # Grand entier (HU, code…) → conserver en texte
                return int(f) if f == int(f) else f
            except ValueError:
                return v              # Texte libre ou date → laisser tel quel

        # Remplacer l'onglet Stock dans main.xlsx
        wb = openpyxl.load_workbook(self.excel_file)
        ws = wb['Stock']

        # Effacer les anciennes données (en-têtes compris)
        ws.delete_rows(1, ws.max_row)

        # Écrire les nouvelles en-têtes et données
        ws.append(list(sap_df.columns))
        for _, row in sap_df.iterrows():
            ws.append([_to_cell(v) for v in row.tolist()])

        self._backup_excel()
        wb.save(self.excel_file)
        print(f"   ✓ Onglet Stock mis à jour dans {self.excel_file}")

        # Cache allégé pour la consultation rapide
        try:
            self._generate_stock_cache(sap_df)
        except Exception as e:
            print(f"   ⚠️  Cache stock non généré : {e}")

        return True

    def load_data(self):
        """Charge toutes les données depuis le fichier Excel"""
        print("📂 Chargement des données...")
        
        # Charger le stock
        # dtype str sur Handling Unit pour éviter la troncature float64 (18+ chiffres)
        self.stock_df = pd.read_excel(
            self.excel_file,
            sheet_name='Stock',
            parse_dates=['Goods Receipt Date', 'Shelf Life Expiration Date'],
            dtype={'Handling Unit': str}
        )

        # Trier par FEFO : date d'expiration d'abord (la plus proche en premier),
        # puis par date de réception pour les articles sans date d'expiration
        self.stock_df = self.stock_df.sort_values(
            ['Shelf Life Expiration Date', 'Goods Receipt Date'],
            na_position='last'  # articles sans date d'expiration pickés en dernier
        )
        
        print(f"   ✓ Stock chargé: {len(self.stock_df)} lignes")
        
        # Charger les besoins de chaque sous-composant
        component_sheets = ['Blade', 'Blade service', 'PCW', 'Upper', 'Lower', 'WEB']
        
        for sheet in component_sheets:
            df = pd.read_excel(self.excel_file, sheet_name=sheet)
            # Renommer les colonnes pour uniformiser
            df.columns = ['Reference', 'Quantity', 'Description', 'Bins', 'Sequence', 'Num_Pal'] + \
                         list(df.columns[6:])
            self.components[sheet] = df[['Reference', 'Quantity', 'Description', 'Sequence', 'Num_Pal']].copy()
            print(f"   ✓ {sheet} chargé: {len(df)} références")
    
    def allocate_stock(self, reference, qty_needed):
        """
        Alloue le stock pour une référence donnée selon FIFO
        
        Args:
            reference: Référence du produit
            qty_needed: Quantité nécessaire
            
        Returns:
            Liste de dictionnaires avec les allocations (emplacement, quantité)
        """
        allocations = []
        remaining_qty = qty_needed
        
        # Filtrer le stock pour cette référence
        stock_items = self.stock_df[self.stock_df['Product'] == reference].copy()
        
        if stock_items.empty:
            return [{
                'storage_bin': 'NON DISPONIBLE',
                'quantity_picked': qty_needed,
                'handling_unit': '',
                'warning': '⚠️ RUPTURE DE STOCK'
            }]
        
        # Allouer selon FIFO (déjà trié par date)
        for idx, row in stock_items.iterrows():
            if remaining_qty <= 0:
                break
            
            available_qty = row['Quantity']
            
            if available_qty > 0:
                qty_to_pick = min(available_qty, remaining_qty)
                
                allocations.append({
                    'storage_bin': row['Storage Bin'],
                    'quantity_picked': qty_to_pick,
                    'handling_unit': row['Handling Unit'],
                    'warning': ''
                })
                
                # Décrémenter le stock
                self.stock_df.at[idx, 'Quantity'] = available_qty - qty_to_pick
                remaining_qty -= qty_to_pick
        
        # Si on n'a pas assez de stock
        if remaining_qty > 0:
            allocations.append({
                'storage_bin': 'MANQUANT',
                'quantity_picked': remaining_qty,
                'handling_unit': '',
                'warning': f'⚠️ MANQUE {remaining_qty} unités'
            })
        
        return allocations
    
    def generate_picking_lists(self, components_filter=None):
        """Génère les picking lists pour les sous-composants sélectionnés.

        Args:
            components_filter: liste de noms de composants à générer.
                               None = tous les composants.
        """
        print("\n📋 Génération des picking lists...")

        for component_name, component_df in self.components.items():
            if components_filter is not None and component_name not in components_filter:
                print(f"   ⏭️  {component_name}: ignoré")
                continue
            picking_data = []
            
            for idx, row in component_df.iterrows():
                reference = row['Reference']
                qty_needed = row['Quantity']
                description = row['Description']
                sequence = row['Sequence']
                num_pal = row['Num_Pal']
                
                # Allouer le stock
                allocations = self.allocate_stock(reference, qty_needed)
                
                # Créer une ligne pour chaque allocation
                for alloc in allocations:
                    picking_data.append({
                        'Séquence': sequence,
                        'N° Palette': num_pal,
                        'Référence': reference,
                        'Description': description,
                        'Quantité': alloc['quantity_picked'],
                        'Emplacement': alloc['storage_bin'],
                        'Handling Unit': alloc['handling_unit'],
                        'Statut': alloc['warning'] if alloc['warning'] else '✓ OK',
                        'Pické': '',
                    })

            # Trier les données par emplacement (allée → travée → niveau)
            picking_data.sort(key=lambda x: sort_location_key(x['Emplacement']))
            
            self.picking_lists[component_name] = pd.DataFrame(picking_data)
            print(f"   ✓ {component_name}: {len(picking_data)} lignes de picking")
    
    def save_picking_lists(self, output_folder='picking_lists', clean_all=False, shared_dir=None):
        """
        Sauvegarde les picking lists dans des fichiers Excel formatés

        Args:
            output_folder: Dossier de sortie pour les picking lists
            clean_all: Si True, supprime tous les PL_*.xlsx avant de générer (workflow complet)
        """
        print(f"\n💾 Sauvegarde des picking lists...")

        # Créer le dossier de sortie
        output_path = Path(output_folder)
        output_path.mkdir(exist_ok=True)

        date_str = datetime.now().strftime('%d-%m-%y')

        if clean_all:
            # Workflow complet : supprimer tous les anciens PL_ pour repartir propre
            old_files = list(output_path.glob("PL_*.xlsx"))
            for f in old_files:
                f.unlink(missing_ok=True)
            if old_files:
                print(f"   🗑️  {len(old_files)} ancienne(s) picking list(s) supprimée(s)")
        else:
            # Étape individuelle : supprimer uniquement les composants régénérés
            removed = 0
            for component_name in self.picking_lists:
                comp_slug = component_name.replace(' ', '_')
                # Compatibilité ancien format (PL_Blade_*.xlsx) + nouveau (#XXXX)
                for old in (list(output_path.glob(f"PL_{comp_slug}_*.xlsx")) +
                            list(output_path.glob(f"PL_#*_{comp_slug}_*.xlsx"))):
                    old.unlink(missing_ok=True)
                    removed += 1
            if removed:
                print(f"   🗑️  {removed} ancienne(s) picking list(s) supprimée(s)")

        for component_name, picking_df in self.picking_lists.items():
            pl_num    = _get_next_pl_number(shared_dir=shared_dir)
            comp_slug = component_name.replace(' ', '_')
            filename  = output_path / f"PL_#{pl_num:04d}_{comp_slug}_{date_str}.xlsx"
            
            # Créer un nouveau classeur Excel avec mise en forme
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # startrow=1 : décale les données en ligne 2 pour laisser la ligne 1 au titre
                picking_df.to_excel(writer, sheet_name='Picking List', index=False, startrow=1)

                workbook = writer.book
                worksheet = writer.sheets['Picking List']

                # Styles communs
                thin_border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'),  bottom=Side(style='thin')
                )
                red_fill   = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
                alt_fill   = PatternFill(start_color='EEEDF5', end_color='EEEDF5', fill_type='solid')
                pique_fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
                body_font  = Font(size=10)
                header_fill = PatternFill(start_color='2B2660', end_color='2B2660', fill_type='solid')
                header_font = Font(color='FFFFFF', bold=True, size=12)

                # Titre (ligne 1) — fusionné sur toutes les colonnes
                num_cols = len(picking_df.columns)
                last_col = openpyxl.utils.get_column_letter(num_cols)
                worksheet.merge_cells(f'A1:{last_col}1')
                title_cell = worksheet['A1']
                title_cell.value = f"PICKING LIST #{pl_num:04d} — {component_name.upper()} — {date_str}"
                title_cell.fill = PatternFill(start_color='1A1545', end_color='1A1545', fill_type='solid')
                title_cell.font = Font(color='FFFFFF', bold=True, size=16)
                title_cell.alignment = Alignment(horizontal='center', vertical='center')
                title_cell.border = thin_border
                worksheet.row_dimensions[1].height = 40

                # En-tête (ligne 2)
                worksheet.row_dimensions[2].height = 28
                for cell in worksheet[2]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border

                # Titre + en-tête répétés sur chaque page imprimée
                worksheet.print_title_rows = '1:2'
                # Figer sous l'en-tête
                worksheet.freeze_panes = 'A3'
                # Filtre automatique sur la ligne d'en-tête uniquement
                worksheet.auto_filter.ref = f"A2:{last_col}{worksheet.max_row}"

                # Mise en page : A4 paysage, colonnes sur une page, hauteur libre
                worksheet.page_setup.paperSize = 9  # A4
                worksheet.page_setup.orientation = 'landscape'
                worksheet.sheet_properties.pageSetUpPr.fitToPage = True
                worksheet.page_setup.fitToWidth = 1
                worksheet.page_setup.fitToHeight = 0
                worksheet.page_setup.copies = 1
                worksheet.page_margins = PageMargins(
                    left=0.25, right=0.25, top=0.5, bottom=0.5,
                    header=0.3, footer=0.3
                )

                # Largeurs de colonnes
                column_widths = {
                    'A': 12,  # Séquence
                    'B': 12,  # N° Palette
                    'C': 16,  # Référence
                    'D': 45,  # Description
                    'E': 10,  # Quantité
                    'F': 22,  # Emplacement
                    'G': 22,  # Handling Unit
                    'H': 22,  # Statut
                    'I': 10,  # Pické
                }
                for col, width in column_widths.items():
                    worksheet.column_dimensions[col].width = width

                # Alignements par colonne
                col_alignments = {
                    'A': Alignment(horizontal='center', vertical='center'),
                    'B': Alignment(horizontal='center', vertical='center'),
                    'C': Alignment(horizontal='center', vertical='center'),
                    'D': Alignment(horizontal='left',   vertical='center', wrap_text=True),
                    'E': Alignment(horizontal='center', vertical='center'),
                    'F': Alignment(horizontal='center', vertical='center'),
                    'G': Alignment(horizontal='center', vertical='center'),
                    'H': Alignment(horizontal='center', vertical='center'),
                    'I': Alignment(horizontal='center', vertical='center'),
                }

                # Lignes de données (démarrent en ligne 3 grâce au décalage startrow=1)
                for row_idx, row in enumerate(worksheet.iter_rows(min_row=3), start=3):
                    worksheet.row_dimensions[row_idx].height = 20
                    is_warning = '⚠️' in str(row[7].value)
                    row_fill = red_fill if is_warning else (alt_fill if row_idx % 2 == 0 else None)
                    for cell in row:
                        cell.font = body_font
                        cell.border = thin_border
                        if cell.column_letter == 'I':
                            cell.fill = pique_fill
                        elif row_fill:
                            cell.fill = row_fill
                        if cell.column_letter in col_alignments:
                            cell.alignment = col_alignments[cell.column_letter]
            
            print(f"   ✓ {filename.name}")
        
        return output_path
    
    def save_updated_stock(self, output_file='stock_updated.xlsx'):
        """
        Sauvegarde le fichier Excel avec le stock mis à jour
        
        Args:
            output_file: Nom du fichier de sortie
        """
        print(f"\n💾 Sauvegarde du stock mis à jour...")
        
        # Charger le classeur original
        wb = openpyxl.load_workbook(self.excel_file)
        
        # Mettre à jour l'onglet Stock
        ws = wb['Stock']
        
        # Effacer les anciennes données (garder les en-têtes)
        ws.delete_rows(2, ws.max_row)
        
        # Écrire les nouvelles données
        # Remplacer les "nan" (str) issus de dtype={'Handling Unit': str} par None
        for idx, row in self.stock_df.iterrows():
            ws.append([None if (isinstance(v, str) and v.lower() == 'nan') else v
                       for v in row.tolist()])
        
        # Sauvegarder
        self._backup_excel(output_file)
        wb.save(output_file)
        print(f"   ✓ Stock mis à jour sauvegardé dans: {output_file}")

        # Régénérer le cache stock pour refléter les quantités après allocation FEFO
        try:
            df_cache = self.stock_df.copy()
            if 'Quantity' in df_cache.columns:
                df_cache = df_cache[
                    pd.to_numeric(df_cache['Quantity'], errors='coerce').fillna(0) > 0
                ]
            self._generate_stock_cache(df_cache)
        except Exception as e:
            print(f"   ⚠️  Cache stock non mis à jour : {e}")

        return output_file
    
    def remove_staging_locations(self, target_file=None):
        """
        Supprime les lignes du stock dont le Storage Bin correspond au pattern
        '8 chiffres + tiret' (ex: 10200650-BLADE) et sauvegarde dans le fichier cible.

        Args:
            target_file: Fichier de destination (par défaut: le fichier source)

        Returns:
            int: Nombre de lignes supprimées
        """
        if target_file is None:
            target_file = self.excel_file

        print(f"\n🗑️  Suppression des emplacements de transit (8 chiffres + tiret)...")

        pattern = re.compile(r'^\d{8}-')

        # Identifier les lignes à supprimer dans le DataFrame en mémoire
        if self.stock_df is not None:
            mask = self.stock_df['Storage Bin'].astype(str).str.match(r'^\d{8}-')
            removed_count = mask.sum()
            self.stock_df = self.stock_df[~mask].reset_index(drop=True)
            print(f"   ✓ {removed_count} lignes supprimées du DataFrame en mémoire")
        else:
            removed_count = 0

        # Mettre à jour directement le fichier Excel
        wb = openpyxl.load_workbook(self.excel_file)
        ws = wb['Stock']

        rows_to_delete = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            storage_bin = str(row[5]) if row[5] is not None else ''
            if pattern.match(storage_bin):
                rows_to_delete.append(row_idx)

        # Supprimer de bas en haut pour ne pas décaler les indices
        for row_idx in reversed(rows_to_delete):
            ws.delete_rows(row_idx)

        self._backup_excel(target_file)
        wb.save(target_file)
        count = len(rows_to_delete)
        print(f"   ✓ {count} lignes supprimées de l'onglet Stock")
        print(f"   ✓ Fichier sauvegardé: {target_file}")

        return count

    def generate_summary_report(self):
        """Génère un rapport récapitulatif"""
        print("\n📊 RAPPORT RÉCAPITULATIF")
        print("=" * 70)
        
        for component_name, picking_df in self.picking_lists.items():
            total_lines = len(picking_df)
            warnings = len(picking_df[picking_df['Statut'].str.contains('⚠️', na=False)])
            ok_lines = total_lines - warnings
            
            print(f"\n{component_name}:")
            print(f"  • Lignes de picking: {total_lines}")
            print(f"  • OK: {ok_lines} ✓")
            print(f"  • Alertes: {warnings} ⚠️")
            
            if warnings > 0:
                print(f"\n  Références en rupture/manquantes:")
                alert_items = picking_df[picking_df['Statut'].str.contains('⚠️', na=False)]
                for _, item in alert_items.iterrows():
                    print(f"    - {item['Référence']}: {item['Statut']}")
        
        print("\n" + "=" * 70)

    def allocate_stock_by_palette(self, reference, num_palettes):
        """
        Alloue N palettes complètes (handling units) pour une référence.
        Chaque ligne du stock = une palette.

        Args:
            reference: Référence du produit
            num_palettes: Nombre de palettes à prendre

        Returns:
            Liste de dictionnaires avec les allocations
        """
        allocations = []
        remaining = num_palettes

        stock_items = self.stock_df[self.stock_df['Product'] == reference].copy()

        if stock_items.empty:
            return [{
                'storage_bin': 'NON DISPONIBLE',
                'quantity_picked': 0,
                'handling_unit': '',
                'warning': '⚠️ RUPTURE DE STOCK'
            }]

        for idx, row in stock_items.iterrows():
            if remaining <= 0:
                break
            if row['Quantity'] > 0:
                allocations.append({
                    'storage_bin': row['Storage Bin'],
                    'quantity_picked': row['Quantity'],
                    'handling_unit': row['Handling Unit'],
                    'warning': ''
                })
                self.stock_df.at[idx, 'Quantity'] = 0
                remaining -= 1

        if remaining > 0:
            allocations.append({
                'storage_bin': 'MANQUANT',
                'quantity_picked': 0,
                'handling_unit': '',
                'warning': f'⚠️ MANQUE {remaining} palette(s)'
            })

        return allocations

    def generate_adhoc_picking_list(self, items, output_folder='picking_lists', shared_dir=None):
        """
        Génère une picking list ad-hoc pour une liste de références avec quantités.

        Args:
            items: liste de dict {'reference': str, 'quantity': float, 'unit': 'palette'|'piece'}
            output_folder: dossier de sortie

        Returns:
            Path du fichier généré
        """
        print("\n🔧 Génération de la picking list ad-hoc...")

        # Charger le stock si pas encore fait
        if self.stock_df is None:
            self.load_data()

        # Construire un dictionnaire de descriptions depuis les composants BOM
        desc_map = {}
        for comp_df in self.components.values():
            for _, row in comp_df.iterrows():
                ref = str(row['Reference']).strip()
                if ref and ref != 'nan':
                    desc_map[ref] = str(row['Description']).strip()

        # Compléter avec les descriptions de l'onglet Stock
        # (pour les références absentes des BOM)
        if self.stock_df is not None:
            desc_col = next(
                (c for c in self.stock_df.columns if 'description' in str(c).lower()),
                None
            )
            if desc_col:
                for _, row in self.stock_df.drop_duplicates(subset='Product').iterrows():
                    ref = str(row['Product']).strip()
                    if ref and ref != 'nan' and ref not in desc_map:
                        val = str(row[desc_col]).strip()
                        if val and val != 'nan':
                            desc_map[ref] = val

        picking_data = []

        for item in items:
            reference = item['reference']
            quantity = item['quantity']
            unit = item['unit']
            description = desc_map.get(reference, '')

            print(f"   → {reference}  ×{quantity:.0f} {unit}(s)")

            if unit == 'palette':
                allocations = self.allocate_stock_by_palette(reference, int(quantity))
            else:
                allocations = self.allocate_stock(reference, quantity)

            for alloc in allocations:
                picking_data.append({
                    'Référence': reference,
                    'Description': description,
                    'Qté demandée': f"{quantity:.0f} {unit}(s)",
                    'Quantité pickée': alloc['quantity_picked'],
                    'Emplacement': alloc['storage_bin'],
                    'Handling Unit': alloc['handling_unit'],
                    'Statut': alloc['warning'] if alloc['warning'] else '✓ OK',
                    'Pické': '',
                })

        # Trier par emplacement
        picking_data.sort(key=lambda x: sort_location_key(x['Emplacement']))

        output_path = Path(output_folder)
        output_path.mkdir(exist_ok=True)

        pl_num   = _get_next_pl_number(shared_dir=shared_dir)
        date_str = datetime.now().strftime('%d-%m-%y_%H%M')
        filename = output_path / f"PL_#{pl_num:04d}_AdHoc_{date_str}.xlsx"

        picking_df = pd.DataFrame(picking_data)

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # startrow=1 : laisse la ligne 1 au titre
            picking_df.to_excel(writer, sheet_name='Picking Ad-Hoc', index=False, startrow=1)

            worksheet = writer.sheets['Picking Ad-Hoc']

            # Styles communs
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'),  bottom=Side(style='thin')
            )
            red_fill   = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
            alt_fill   = PatternFill(start_color='EEEDF5', end_color='EEEDF5', fill_type='solid')
            pique_fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
            body_font  = Font(size=10)
            header_fill = PatternFill(start_color='2B2660', end_color='2B2660', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True, size=12)

            # Titre (ligne 1) — fusionné sur toutes les colonnes
            num_cols = len(picking_df.columns)
            last_col = openpyxl.utils.get_column_letter(num_cols)
            worksheet.merge_cells(f'A1:{last_col}1')
            title_cell = worksheet['A1']
            title_cell.value = f"PICKING LIST AD-HOC #{pl_num:04d} — {date_str}"
            title_cell.fill = PatternFill(start_color='1A1545', end_color='1A1545', fill_type='solid')
            title_cell.font = Font(color='FFFFFF', bold=True, size=16)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            title_cell.border = thin_border
            worksheet.row_dimensions[1].height = 40

            # En-tête (ligne 2)
            worksheet.row_dimensions[2].height = 28
            for cell in worksheet[2]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

            # Titre + en-tête répétés sur chaque page imprimée
            worksheet.print_title_rows = '1:2'
            # Figer sous l'en-tête
            worksheet.freeze_panes = 'A3'
            # Filtre automatique sur la ligne d'en-tête uniquement
            worksheet.auto_filter.ref = f"A2:{last_col}{worksheet.max_row}"

            # Mise en page : A4 paysage, colonnes sur une page, hauteur libre
            worksheet.page_setup.paperSize = 9  # A4
            worksheet.page_setup.orientation = 'landscape'
            worksheet.sheet_properties.pageSetUpPr.fitToPage = True
            worksheet.page_setup.fitToWidth = 1
            worksheet.page_setup.fitToHeight = 0
            worksheet.page_setup.copies = 1
            worksheet.page_margins = PageMargins(
                left=0.25, right=0.25, top=0.5, bottom=0.5,
                header=0.3, footer=0.3
            )

            # Largeurs de colonnes
            column_widths = {
                'A': 18,  # Référence
                'B': 45,  # Description
                'C': 18,  # Qté demandée
                'D': 14,  # Quantité pickée
                'E': 22,  # Emplacement
                'F': 22,  # Handling Unit
                'G': 22,  # Statut
                'H': 10,  # Pické
            }
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width

            # Alignements par colonne
            col_alignments = {
                'A': Alignment(horizontal='center', vertical='center'),
                'B': Alignment(horizontal='left',   vertical='center', wrap_text=True),
                'C': Alignment(horizontal='center', vertical='center'),
                'D': Alignment(horizontal='center', vertical='center'),
                'E': Alignment(horizontal='center', vertical='center'),
                'F': Alignment(horizontal='center', vertical='center'),
                'G': Alignment(horizontal='center', vertical='center'),
                'H': Alignment(horizontal='center', vertical='center'),
            }

            # Lignes de données (démarrent en ligne 3 grâce au décalage startrow=1)
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=3), start=3):
                worksheet.row_dimensions[row_idx].height = 20
                is_warning = '⚠️' in str(row[6].value)
                row_fill = red_fill if is_warning else (alt_fill if row_idx % 2 == 0 else None)
                for cell in row:
                    cell.font = body_font
                    cell.border = thin_border
                    if cell.column_letter == 'H':
                        cell.fill = pique_fill
                    elif row_fill:
                        cell.fill = row_fill
                    if cell.column_letter in col_alignments:
                        cell.alignment = col_alignments[cell.column_letter]

        print(f"   ✓ {filename.name} — {len(picking_data)} ligne(s) de picking")
        return filename

    def generate_bom_sheet(self):
        """
        Génère/met à jour l'onglet 'BOM' dans main.xlsx.
        Vue plate de toutes les références de tous les sous-composants,
        avec une ligne par référence et une colonne 'Composant'.
        """
        print("\n📑 Génération de l'onglet BOM consolidé...")

        component_sheets = ['Blade', 'Blade service', 'PCW', 'Upper', 'Lower', 'WEB']
        component_colors = {
            'Blade':         'E3F0F8',
            'Blade service': 'E3F8EC',
            'PCW':           'FFF8E3',
            'Upper':         'F8E3F8',
            'Lower':         'F8EFE3',
            'WEB':           'E3F8F8',
        }

        all_rows = []
        for sheet in component_sheets:
            try:
                df = pd.read_excel(self.excel_file, sheet_name=sheet)
            except Exception:
                print(f"   ⚠️  Onglet '{sheet}' introuvable — ignoré")
                continue
            count = 0
            for _, row in df.iterrows():
                ref = str(row.iloc[0]).strip()
                if not ref or ref == 'nan':
                    continue
                qty = row.iloc[1]
                des = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else ''
                all_rows.append({
                    'Composant':   sheet,
                    'Référence':   ref,
                    'Description': des,
                    'Quantité':    float(qty) if pd.notna(qty) else 0,
                })
                count += 1
            print(f"   • {sheet} : {count} référence(s)")

        wb = openpyxl.load_workbook(self.excel_file)

        # Supprimer l'onglet BOM s'il existe déjà
        if 'BOM' in wb.sheetnames:
            del wb['BOM']
        ws = wb.create_sheet('BOM')

        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'),  bottom=Side(style='thin')
        )

        # En-têtes
        headers = ['Composant', 'Référence', 'Description', 'Quantité']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = PatternFill(start_color='003755', end_color='003755', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # Données
        for row_idx, row_data in enumerate(all_rows, 2):
            color = component_colors.get(row_data['Composant'], 'FFFFFF')
            fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            values = [row_data['Composant'], row_data['Référence'],
                      row_data['Description'], row_data['Quantité']]
            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = fill
                cell.border = thin_border
                if col_idx == 4:
                    cell.alignment = Alignment(horizontal='center')

        # Largeurs de colonnes
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 55
        ws.column_dimensions['D'].width = 12

        self._backup_excel()
        wb.save(self.excel_file)
        print(f"   ✓ Onglet BOM mis à jour — {len(all_rows)} lignes au total")
        return len(all_rows)


def archive_picking_lists(picking_folder='picking_lists', shared_dir=None):
    """
    Déplace les PL_*.xlsx du dossier picking_lists vers une archive numérotée.

    Returns:
        dict: Entrée d'archive créée, ou None si aucun fichier à archiver.
    """
    base         = Path(shared_dir) if shared_dir else Path('.')
    archive_dir  = base / 'picking_lists_archive'
    index_file   = archive_dir / 'archive_index.json'
    picking_path = Path(picking_folder)
    files = sorted(picking_path.glob('PL_*.xlsx'))

    if not files:
        print("⚠️  Aucune picking list à archiver.")
        return None

    archive_dir.mkdir(parents=True, exist_ok=True)

    index = {'next_number': 1, 'archives': []}
    if index_file.exists():
        try:
            index = json.loads(index_file.read_text(encoding='utf-8'))
        except Exception:
            pass

    num         = index['next_number']
    now         = datetime.now()
    date_str    = now.strftime('%Y-%m-%d')
    folder_name = f"#{num:04d}_{date_str}"
    folder_path = archive_dir / folder_name
    folder_path.mkdir(exist_ok=True)

    print(f"\n📦 Archivage → {folder_name}")
    moved = []
    for f in files:
        shutil.move(str(f), str(folder_path / f.name))
        moved.append(f.name)
        print(f"   • {f.name}")

    components = []
    for fname in moved:
        parts = fname.replace('.xlsx', '').split('_')
        if len(parts) >= 3:
            comp = ' '.join(p for p in parts[1:-1] if not p.startswith('#'))
            if comp and comp not in components:
                components.append(comp)

    entry = {
        'number':     num,
        'date':       date_str,
        'datetime':   now.strftime('%d/%m/%Y %H:%M'),
        'folder':     folder_name,
        'files':      moved,
        'components': components,
    }
    index['archives'].append(entry)
    index['next_number'] = num + 1

    index_file.write_text(json.dumps(index, indent=2, ensure_ascii=False), encoding='utf-8')
    print(f"   ✅ Archive #{num:04d} créée ({len(moved)} fichier(s))")
    return entry


def load_archive_index(shared_dir=None):
    """Charge l'index des archives de picking lists."""
    base        = Path(shared_dir) if shared_dir else Path('.')
    archive_dir = base / 'picking_lists_archive'
    index_file  = archive_dir / 'archive_index.json'
    archive_dir.mkdir(parents=True, exist_ok=True)
    if index_file.exists():
        try:
            return json.loads(index_file.read_text(encoding='utf-8'))
        except Exception:
            pass
    return {'next_number': 1, 'archives': []}


def clean_staging_stock(excel_file='main.xlsx'):
    """
    Fonction standalone : supprime les lignes de stock avec un Storage Bin
    au format '8 chiffres + tiret' (ex: 10200650-BLADE) et sauvegarde dans le fichier source.

    Peut être appelée indépendamment sans générer de picking lists.

    Args:
        excel_file: Chemin vers le fichier Excel principal (défaut: main.xlsx)
    """
    print("=" * 70)
    print("  NETTOYAGE DU STOCK - SUPPRESSION EMPLACEMENTS DE TRANSIT")
    print("=" * 70)

    generator = PickingListGenerator(excel_file)
    generator.load_data()
    count = generator.remove_staging_locations(target_file=excel_file)

    print(f"\n✅ TERMINÉ! {count} lignes supprimées dans {excel_file}")
    return count


def main():
    """Fonction principale"""
    print("=" * 70)
    print("  GÉNÉRATEUR DE PICKING LISTS - BLADE D'ÉOLIENNES")
    print("=" * 70)
    
    # Chemin du fichier Excel
    excel_file = 'main.xlsx'
    
    # Créer le générateur
    generator = PickingListGenerator(excel_file)

    # Importer le stock depuis l'export SAP
    generator.import_stock_from_sap()

    # Charger les données
    generator.load_data()

    # Supprimer les emplacements de transit (8 chiffres + tiret) avant allocation
    generator.remove_staging_locations(target_file=excel_file)

    # Générer les picking lists
    generator.generate_picking_lists()
    
    # Sauvegarder les picking lists
    output_folder = generator.save_picking_lists(output_folder='picking_lists')

    # Sauvegarder le stock mis à jour directement dans main.xlsx
    generator.save_updated_stock(output_file=excel_file)

    # Afficher le rapport
    generator.generate_summary_report()

    print(f"\n✅ TERMINÉ!")
    print(f"\n📁 Fichiers générés:")
    print(f"   • Picking lists: {output_folder}/")
    
    return generator

if __name__ == "__main__":
    generator = main()
